import pandas as pd
import math
import os
import numpy as np
import openpyxl
from openpyxl import Workbook

path = os.getcwd()
files = os.listdir(path)
files_csv = [f for f in files if f[-3:] == 'csv']
print('number of files: '+str(len(files_csv)))
for work_item in range (0,len(files_csv)):
    IO = files_csv[work_item]
    #print(IO)
    # 輸入的檔案名稱，需先轉成 xls 檔，並下移一行
    #IO = 'SAA_table_all.csv'
    sheet = pd.read_csv(IO,header=None,sep=",")
    # 創建一個空白活頁簿物件
    wb = Workbook()
    # 選取正在工作中的表單
    ws = wb.active

    # Store 24 Varient Name into varienNames list
    varientNames = ['SAA2a (RS-)', 'SAA2b (RS-)', 'SAA1g (RS-)', 'SAAU2 (RS-)', \
               'SAA1a (RS-)', 'SAAU1 (RS-)', 'SAA2a (R-)', 'SAA2b (R-)', \
                'SAA1b (RS-)', 'SAA1g (R-)', 'SAAU3 (RS-)', 'SAAU2 (R-)', \
                'SAA1a (R-)', 'SAAU1 (R-)', 'SAA1b (R-)', 'SAAU3 (R-)', \
                'SAA2a', 'SAA2b', 'SAA1g', 'SAAU2', 'SAA1a', 'SAAU1', 'SAA1b', 'SAAU3']
    binNames = []
    # Store 86 Bin Name into binNames list
    for i in range(86):
        binNames.append('bin'+str(i+1))
    ISD_1 = []
    ISD_2 = []
    MAD = []
    print('file maximum rows: '+str(len(sheet.values)))
    ws.cell(row = 1, column = 1, value = "test")
    print('---------- assign title and initialize value ----------')
    # assign title and initialize value
    for i in range(1,len(sheet.values)+1):
        if ((i%3) == 1):
            for j in range (0,4):
                ws.cell(row = i+1, column = j+1, value = sheet.values[i-1][j])
    for i in range(1,len(sheet.values)+1):
        for j in range(0,110):
            if((i%3) == 1):
                if(j<24):
                    ws.cell(row = i+1, column = j+5, value = varientNames[j])
                else:
                    ws.cell(row = i+1, column = j+5, value = binNames[j-24])
            elif((i%3) != 1):
                if(j<24):
                    ws.cell(row = i+1, column = j+5, value = 0)
                else:
                    ws.cell(row = i+1, column = j+5, value = 0)
        print('processing row:'+ str(i))
        
    print('---------- start finding match value ----------')
    row = 0
    ISD_length = 0 # 'ISD' col length
    for i in range(0,len(sheet.values),3):
        q_num = 0 # '?' number
        print('processing row: '+ str(i))
        for j in range(4,len(sheet.values[1])):
            varient_num = 0
            bin_num = 0
            if(sheet.values[i][j] == '?'):
                ws.cell(row = i+2, column = q_num+115, value = '?') # 115 = 4(first 4 metadata) + 24(varient_num) + 86(bin_num) + 1(this col)
                ws.cell(row = i+3, column = q_num+115, value = sheet.values[i+1][j])
                ws.cell(row = i+4, column = q_num+115, value = sheet.values[i+2][j])
                q_num = q_num+1
            elif(sheet.values[i][j] == 'ISD:'):
                if(ISD_length < q_num+116):
                    ISD_length = q_num+116
                ISD_1.append(sheet.values[i+1][j]) 
                ISD_2.append(sheet.values[i+2][j])
            elif(sheet.values[i][j] == 'MAD:'):
                MAD.append(sheet.values[i+1][j])
                break
            else:
                for k in range(varient_num,24):
                    if(sheet.values[i][j] == varientNames[k]):
                        ws.cell(row = i+3, column = k+5, value = sheet.values[i+1][j])
                        ws.cell(row = i+4, column = k+5, value = sheet.values[i+2][j])
                        varient_num = k
                        break;
                if(k != varient_num): # if find varient_num before, no need to check bin_num
                    for k in range(bin_num,86):
                        if(sheet.values[i][j] == binNames[k]):
                            ws.cell(row = i+3, column = k+29, value = sheet.values[i+1][j])
                            ws.cell(row = i+4, column = k+29, value = sheet.values[i+2][j])
                            bin_num = k
                            break;

        row = row+1
    i = 0
    #print(ISD_length)
    # 將 ISD 跟 MSD 值寫入
    for j in range(0,len(sheet.values),3):
        ws.cell(row = j+2, column = ISD_length, value = 'ISD:')
        ws.cell(row = j+3, column = ISD_length, value = ISD_1[i])
        ws.cell(row = j+4, column = ISD_length, value = ISD_2[i])
        ws.cell(row = j+2, column = ISD_length+2, value = 'MAD:')
        ws.cell(row = j+3, column = ISD_length+2, value = MAD[i])
        i = i+1
    print('file number '+str(work_item+1) +' is finished')
    # 儲存成 create_sample.xls 檔案
    wb.save('tmp.xls')
    data_xls = pd.read_excel('tmp.xls',index_col=None)
    if('.csv' in files_csv[work_item]):
        files_csv[work_item] = files_csv[work_item][:-4]
    data_xls.to_csv(files_csv[work_item]+'_排列.csv', encoding='utf-8',sep=',',index=False,header=None)
    os.remove('tmp.xls')
    work_item = work_item + 1