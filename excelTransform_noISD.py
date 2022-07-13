import pandas as pd
import math
import os
import numpy as np
import openpyxl
from openpyxl import Workbook

path = os.getcwd()
files = os.listdir(path)
files_csv = [f for f in files if f[-7:] == '_P.xlsx']
print('number of files: '+str(len(files_csv)))
for work_item in range (0,len(files_csv)):
    IO = files_csv[work_item]
    #IO = 'SAA_table_all.csv'
    sheet = pd.read_excel(IO,header=None)
    # 創建一個空白活頁簿物件
    wb = Workbook()
    # 選取正在工作中的表單
    ws = wb.active

    # Store 24 Varient Name into varienNames list
    varientNames = ['SAA2a (RS-)', 'SAA2b (RS-)', 'SAA1g (RS-)', 'SAAU2 (RS-)', \
               'SAA1a (RS-)', 'SAAU1 (RS-)', 'SAA2a (R-)', 'SAA2b (R-)', \
                'SAA1b (RS-)', 'SAA1g (R-)', 'SAAU3 (RS-)', 'SAAU2 (R-)', \
                'SAA1a (R-)', 'SAAU1 (R-)', 'SAA1b (R-)', 'SAAU3 (R-)', \
                'SAA2a', 'SAA2b', 'SAA1g', 'SAAU2', 'SAA1a', 'SAAU1', 'SAA1b', 'SAAU3']
    binNames = []
    # Store 86 Bin Name into binNames list
    for i in range(86):
        binNames.append('bin'+str(i+1))
    ISD_1 = []
    ISD_2 = []
    MAD = []
    print('file maximum rows: '+str(len(sheet.values)))
    ws.cell(row = 1, column = 1, value = "test")
    # assign title and initialize value
    #for i in range(1,len(sheet.values)+1):
    #    print('processing row:'+ str(i))
    #    if ((i%3) == 1):
    #        for j in range (0,4):
    #            ws.cell(row = i+1, column = j+1, value = sheet.values[i-1][j])
    
    '''
    for j in range(0,110):
        if(j<24):
            ws.cell(row = 2, column = j+2, value = varientNames[j])
        else:
            ws.cell(row = 2, column = j+2, value = binNames[j-24])
    ws.cell(row = 2, column = 112, value = 'ISD')
    ws.cell(row = 2, column = 113, value = 'MAD1')
    ws.cell(row = 2, column = 114, value = 'MAD2')
    '''
    # assign title
    title_row=2
    for i in range(2,len(sheet.values)+1):
        ws.cell(row = title_row, column = 1, value = sheet.values[i-1][0])
        title_row=title_row+3
    # find existed peak
    row_num=2
    for i in range(2,len(sheet.values)+1):
        j = 0
        col_num = 2
        # find existed peak and bin
        for j in range(1,110):
            if (sheet.values[i-1][j] != 0):
                if (j < 25):
                    ws.cell(row = row_num, column = col_num, value = varientNames[j-1])
                else:
                    ws.cell(row = row_num, column = col_num, value = binNames[j-24-1])
                ws.cell(row = row_num+2, column = col_num, value = sheet.values[i-1][j])
                col_num = col_num + 1
        # find existed ISD and MAD
        col_num = col_num + 1
        ws.cell(row = row_num, column = col_num, value = 'ISD')
        ws.cell(row = row_num + 2, column = col_num, value = sheet.values[i-1][112])
        col_num = col_num + 1
        ws.cell(row = row_num, column = col_num, value = 'MAD1')
        ws.cell(row = row_num + 2, column = col_num, value = sheet.values[i-1][113])
        col_num = col_num + 1
        ws.cell(row = row_num, column = col_num, value = 'MAD2')
        ws.cell(row = row_num + 2, column = col_num, value = sheet.values[i-1][114])
        row_num = row_num + 3
    i = 0
    print('file number '+str(work_item+1) +' is finished')
    # 儲存成 create_sample.xls 檔案
    wb.save('tmp.xls')
    data_xls = pd.read_excel('tmp.xls',index_col=None)
    if('_P.xlsx' in files_csv[work_item]):
        files_csv[work_item] = files_csv[work_item][:-8]
    data_xls.to_csv(files_csv[work_item]+'.csv', encoding='utf-8',sep=',',index=False,header=None)
    os.remove('tmp.xls')
    work_item = work_item + 1