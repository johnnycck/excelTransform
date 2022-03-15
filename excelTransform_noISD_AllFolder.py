import pandas as pd
import math
import os
import numpy as np
import openpyxl
from openpyxl import Workbook

path = os.getcwd()
for root, dirs, files in os.walk(path):
    files_csv = [f for f in files if f[-3:] == 'csv']
    files_csv = [f for f in files_csv if f[:6] != '【Test】']
    files_csv = [f for f in files_csv if f[:7] != '【Train】']
    #files_csv = [f for f in files_csv if f[:6] == 'Passed' or if f[:3] == 'ISD' ]
    #files_csv = [f for f in files_csv if f[:6] == 'Passed']
        
    if(len(files_csv) > 0):
        print('number of files: '+str(len(files_csv)))
        for work_item in range (0,len(files_csv)):
            IO = (os.path.join(root, files_csv[work_item]))
            print(IO)
            #IO = 'SAA_table_all.csv'
            sheet = pd.read_csv(IO,header=None,sep=",")
            # 創建一個空白活頁簿物件
            wb = Workbook()
            # 選取正在工作中的表單
            ws = wb.active

            # Store 24 Varient Name into varienNames list
            varientNames = ['SAA2a (RS-)', 'SAA2b (RS-)', 'SAA1g (RS-)', 'SAAU2 (RS-)', \
                            'SAA1a (RS-)', 'SAAU1 (RS-)', 'SAA2a (R-)', 'SAA2b (R-)', \
                            'SAA1b (RS-)', 'SAA1g (R-)', 'SAAU3 (RS-)', 'SAAU2 (R-)', \
                            'SAA1a (R-)', 'SAAU1 (R-)', 'SAA1b (R-)', 'SAAU3 (R-)', \
                            'SAA2a', 'SAA2b', 'SAA1g', 'SAAU2', 'SAA1a', 'SAAU1', 'SAA1b', 'SAAU3']
            binNames = []
            # Store 86 Bin Name into binNames list
            for i in range(86):
                binNames.append('bin'+str(i+1))
            ISD_1 = []
            ISD_2 = []
            MAD = []
            print('file maximum rows: '+str(len(sheet.values)))
            ws.cell(row = 1, column = 1, value = "test")
            print('---------- assign title and initialize value ----------')
            # assign title and initialize value
            #for i in range(1,len(sheet.values)+1):
            #    print('processing row:'+ str(i))
            #    if ((i%3) == 1):
            #        for j in range (0,4):
            #            ws.cell(row = i+1, column = j+1, value = sheet.values[i-1][j])
            title_row=3
            for j in range(0,110):
                if(j<24):
                    ws.cell(row = 2, column = j+2, value = varientNames[j])
                else:
                    ws.cell(row = 2, column = j+2, value = binNames[j-24])
            for i in range(1,len(sheet.values)+1):
                #print('processing row:'+ str(i))
                for j in range(0,110):
                    if((i%3) == 1):
                        if(j<1):
                            ws.cell(row = title_row, column = j+1, value = sheet.values[i-1][j])
                    elif((i%3) == 0):
                        if(j<109):
                            ws.cell(row = title_row, column = j+2, value = 0)
                        else:
                            ws.cell(row = title_row, column = j+2, value = 0)
                            title_row=title_row+1
           
            print('---------- start finding match value ----------')
            row = 0
            content_row=3
            for i in range(0,len(sheet.values),3):
                #print('processing row: '+ str(i))
                for j in range(1,len(sheet.values[1])):
                    varient_num = 0
                    bin_num = 0
                    for k in range(varient_num,24):
                        if(sheet.values[i][j] == varientNames[k]):
                            #ws.cell(row = i+3, column = k+2, value = sheet.values[i+1][j])
                            ws.cell(row = content_row, column = k+2, value = sheet.values[i+2][j])
                            varient_num = k
                            break;
                    if(k != varient_num): # if find varient_num before, no need to check bin_num
                        for k in range(bin_num,86):
                            if(sheet.values[i][j] == binNames[k]):
                                #ws.cell(row = i+3, column = k+26, value = sheet.values[i+1][j])
                                ws.cell(row = content_row, column = k+26, value = sheet.values[i+2][j])
                                bin_num = k
                                break;
                row = row+1
                content_row=content_row+1
            i = 0
            print('file number '+str(work_item+1) +' is finished')
            # 儲存成 create_sample.xls 檔案
            wb.save('tmp.xls')
            data_xls = pd.read_excel('tmp.xls',index_col=None)
            #if('.csv' in files_csv[work_item]):
            #    files_csv[work_item] = files_csv[work_item][:-4]
            if('.csv' in IO):
                IO = IO[:-4]
            csvName = 'tmp'
            if('AllPeaks' in IO):
                csvName = 'SAA_table_all'
            elif('FilterPeak' in IO):
                csvName = 'SAA_table_filt'
            data_xls.to_csv(root+'/'+csvName+'_排列.csv', encoding='utf-8',sep=',',index=False,header=None)
            #data_xls.to_csv(IO+'_排列.csv', encoding='utf-8',sep=',',index=False,header=None)
            os.remove('tmp.xls')
            work_item = work_item + 1